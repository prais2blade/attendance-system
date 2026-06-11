from django.shortcuts import (
    render,
    get_object_or_404
)
from django.utils import timezone
from django.http import FileResponse
from django.conf import settings
import os

from apps.attendance.models import Attendance
from .models import Student, Parent, StudentParent
from io import BytesIO

from django.http import FileResponse

from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
import secrets

from openpyxl import load_workbook

from django.contrib import messages
from django.shortcuts import redirect

from .forms import StudentImportForm
from django.db.models import Q
from django.contrib import messages
from django.shortcuts import redirect

from .forms import StudentForm
import zipfile

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter



# Functions for operations on students

def student_detail(request, pk):

    student = get_object_or_404(
        Student,
        pk=pk
    )

    attendance_history = (
        Attendance.objects
        .filter(student=student)
        .order_by("-date")
    )

    context = {
        "student": student,
        "attendance_history": attendance_history
    }

    return render(
        request,
        "attendance/detail.html",
        context
    )
    
    
def student_id_card(request, pk):

    student = get_object_or_404(
        Student,
        pk=pk
    )

    buffer = BytesIO()

    pdf = canvas.Canvas(
        buffer,
        pagesize=(85.6 * mm, 54 * mm)
    )

    pdf.setStrokeColor(colors.black)

    pdf.rect(
        5,
        5,
        230,
        140
    )

    pdf.setFont(
        "Helvetica-Bold",
        12
    )

    pdf.drawString(
        15,
        130,
        "CODECAMP INNOVATION HUB"
    )

    pdf.setFont(
        "Helvetica",
        9
    )

    pdf.drawString(
        15,
        110,
        f"Name: {student.first_name} {student.last_name}"
    )

    pdf.drawString(
        15,
        95,
        f"ID: {student.student_id}"
    )
    
    if student.qr_code:
        pdf.drawImage(
            student.qr_code.path,
            160,
            40,
            width=60,
            height=60
        )
    if student.photo:
        pdf.drawImage(
            student.photo.path,
            15,
            40,
            width=50,
            height=50
        )

    pdf.save()

    buffer.seek(0)

    return FileResponse(
        buffer,
        as_attachment=True,
        filename=f"{student.student_id}.pdf"
    )
    
def parent_portal(request, token):

    student = get_object_or_404(
        Student,
        portal_token=token
    )

    today = timezone.localdate()

    attendance = (
        Attendance.objects
        .filter(
            student=student,
            date=today
        )
        .first()
    )

    history = (
        Attendance.objects
        .filter(student=student)
        .order_by("-date")[:30]
    )

    context = {
        "student": student,
        "attendance": attendance,
        "history": history
    }

    return render(
        request,
        "students/parent_portal.html",
        context
    )
    

def import_students(request):

    form = StudentImportForm()

    if request.method == "POST":

        form = StudentImportForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            excel_file = request.FILES[
                "excel_file"
            ]

            workbook = load_workbook(
                excel_file
            )

            sheet = workbook.active

            count = 0

            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                first_name = row[0]
                last_name = row[1]
                date_of_birth = row[2]
                gender = row[3]
                class_name = row[4]

                parent_name = row[5]
                parent_email = row[6]
                parent_phone = row[7]
                parent_whatsapp = row[8]

                relationship = (
                    row[9]
                    if len(row) > 9 and row[9]
                    else "Guardian"
                )

                if not first_name:
                    continue

                student = Student.objects.create(

                    first_name=first_name,

                    last_name=last_name or "",

                    date_of_birth=date_of_birth,

                    gender=gender or "",

                    class_name=class_name or "",

                    portal_token=secrets.token_urlsafe(
                        16
                    )

                )

                if parent_name:

                    parent, created = Parent.objects.get_or_create(

                        full_name=parent_name,

                        defaults={

                            "email": parent_email,

                            "phone_number": parent_phone or "",

                            "whatsapp_number": parent_whatsapp or ""

                        }

                    )

                    StudentParent.objects.get_or_create(

                        student=student,

                        parent=parent,

                        defaults={

                            "relationship": relationship

                        }

                    )

                count += 1

            messages.success(

                request,

                f"{count} students imported successfully."

            )

            return redirect(

                "student_list"

            )

    context = {

        "form": form

    }

    return render(

        request,

        "students/import.html",

        context

    )
    

def student_list(request):

    search = request.GET.get(
        "search",
        ""
    )

    students = Student.objects.all().order_by(
        "-id"
    )

    if search:

        students = students.filter(

            Q(student_id__icontains=search)

            |

            Q(first_name__icontains=search)

            |

            Q(last_name__icontains=search)

        )

    context = {

        "students": students,

        "search": search

    }

    return render(

        request,

        "students/list.html",

        context

    )
    
def create_student(request):

    form = StudentForm()

    if request.method == "POST":

        form = StudentForm(

            request.POST,

            request.FILES

        )

        if form.is_valid():

            parent_name = form.cleaned_data.get(
                "parent_name"
            )

            parent_email = form.cleaned_data.get(
                "parent_email"
            )

            parent_phone = form.cleaned_data.get(
                "parent_phone"
            )

            parent_whatsapp = form.cleaned_data.get(
                "parent_whatsapp"
            )

            relationship = form.cleaned_data.get(
                "relationship"
            )

            student = form.save(
                commit=False
            )

            student.save()

            if parent_name:

                parent, created = Parent.objects.get_or_create(

                    full_name=parent_name,

                    defaults={

                        "email": parent_email,

                        "phone_number": parent_phone,

                        "whatsapp_number": parent_whatsapp,

                    }

                )

                StudentParent.objects.get_or_create(

                    student=student,

                    parent=parent,

                    defaults={

                        "relationship": relationship

                    }

                )

            messages.success(

                request,

                f"{student.first_name} created successfully."

            )

            return redirect(

                "student_detail",

                pk=student.pk

            )

    context = {

        "form": form

    }

    return render(

        request,

        "students/create.html",

        context

    )
    

def edit_student(request, pk):

    student = get_object_or_404(
        Student,
        pk=pk
    )

    form = StudentForm(
        instance=student
    )

    if request.method == "POST":

        form = StudentForm(
            request.POST,
            request.FILES,
            instance=student
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Student updated successfully."
            )

            return redirect(
                "student_detail",
                pk=student.pk
            )

    context = {

        "form": form,

        "student": student

    }

    return render(
        request,
        "students/edit.html",
        context
    )
    
def delete_student(request, pk):

    student = get_object_or_404(
        Student,
        pk=pk
    )

    if request.method == "POST":

        student_name = (
            f"{student.first_name} "
            f"{student.last_name}"
        )

        student.delete()

        messages.success(
            request,
            f"{student_name} deleted successfully."
        )

        return redirect(
            "student_list"
        )

    context = {
        "student": student
    }

    return render(
        request,
        "students/delete.html",
        context
    )
    
def bulk_qr_download(request):

    student_ids = request.POST.getlist(
        "students"
    )

    students = Student.objects.filter(
        id__in=student_ids
    )

    buffer = BytesIO()

    zip_file = zipfile.ZipFile(
        buffer,
        "w"
    )

    for student in students:

        if student.qr_code:

            zip_file.write(

                student.qr_code.path,

                arcname=f"{student.student_id}.png"

            )

    zip_file.close()

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/zip"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="qrcodes.zip"'

    return response
    
    
def bulk_id_cards(request):

    if request.method != "POST":

        return redirect("student_list")

    student_ids = request.POST.getlist(
        "students"
    )

    students = Student.objects.filter(
        id__in=student_ids
    )

    buffer = BytesIO()

    pdf = canvas.Canvas(buffer)

    for student in students:

        pdf.setPageSize(
            (85.6 * mm, 54 * mm)
        )

        pdf.setStrokeColor(colors.black)

        pdf.rect(
            5,
            5,
            230,
            140
        )

        pdf.setFont(
            "Helvetica-Bold",
            12
        )

        pdf.drawString(
            15,
            130,
            "CODECAMP INNOVATION HUB"
        )

        pdf.setFont(
            "Helvetica",
            9
        )

        pdf.drawString(
            15,
            110,
            f"Name: {student.first_name} {student.last_name}"
        )

        pdf.drawString(
            15,
            95,
            f"ID: {student.student_id}"
        )

        if student.qr_code:

            pdf.drawImage(
                student.qr_code.path,
                160,
                40,
                width=60,
                height=60
            )

        if student.photo:

            pdf.drawImage(
                student.photo.path,
                15,
                40,
                width=50,
                height=50
            )

        pdf.showPage()

    pdf.save()

    buffer.seek(0)

    return FileResponse(
        buffer,
        as_attachment=True,
        filename="student_id_cards.pdf"
    )
    
def download_student_template(request):

    file_path = os.path.join(

        settings.BASE_DIR,

        "static",

        "templates",

        "student_import_template.xlsx"

    )

    return FileResponse(

        open(file_path, "rb"),

        as_attachment=True,

        filename="student_import_template.xlsx"

    )


def download_student_template(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Students"

    headers = [

        "first_name",
        "last_name",
        "date_of_birth",
        "gender",
        "class_name",
        "parent_name",
        "parent_email",
        "parent_phone",
        "parent_whatsapp",
        "relationship",

    ]

    header_fill = PatternFill(

        start_color="1E40AF",
        end_color="1E40AF",
        fill_type="solid"

    )

    header_font = Font(

        bold=True,
        color="FFFFFF"

    )

    for col_num, header in enumerate(

        headers,

        start=1

    ):

        cell = sheet.cell(

            row=1,

            column=col_num

        )

        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

    sample_row = [

        "John",
        "Doe",
        "2015-04-20",
        "Male",
        "JSS1",
        "Jane Doe",
        "jane@example.com",
        "08012345678",
        "08012345678",
        "Mother",

    ]

    for col_num, value in enumerate(

        sample_row,

        start=1

    ):

        sheet.cell(

            row=2,

            column=col_num

        ).value = value

    for column in sheet.columns:

        max_length = 0

        column_letter = get_column_letter(

            column[0].column

        )

        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(

                        str(cell.value)

                    )

            except:

                pass

        sheet.column_dimensions[
            column_letter
        ].width = max_length + 5

    sheet.freeze_panes = "A2"

    instructions = workbook.create_sheet(

        title="Instructions"

    )

    instructions["A1"] = (
        "CODECAMP ATTENDANCE SYSTEM "
        "STUDENT IMPORT TEMPLATE"
    )

    instructions["A1"].font = Font(

        bold=True,
        size=14

    )

    instructions["A3"] = (
        "Fill one student per row."
    )

    instructions["A4"] = (
        "Do not change the header names."
    )

    instructions["A5"] = (
        "Date format must be YYYY-MM-DD."
    )

    instructions["A6"] = (
        "Gender examples: Male, Female."
    )

    instructions["A7"] = (
        "Relationship examples: "
        "Father, Mother, Guardian."
    )

    instructions["A8"] = (
        "Parent email is optional."
    )

    instructions["A9"] = (
        "Parent phone numbers should "
        "include leading zero."
    )

    response = HttpResponse(

        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )

    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="student_import_template.xlsx"'
    )

    workbook.save(response)

    return response