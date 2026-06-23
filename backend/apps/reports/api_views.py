from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response

from apps.students.models import Student, StudentParent
from apps.attendance.models import Attendance
from django.db.models import Count
from django.http import HttpResponse
from openpyxl import Workbook
import csv


@api_view(["GET"])
def report_summary_api(request):

    today = timezone.now().date()

    total_students = Student.objects.count()

    present_today = Attendance.objects.filter(
        date=today
    ).count()

    checked_out = Attendance.objects.filter(
        date=today,
        check_out__isnull=False
    ).count()

    in_center = Attendance.objects.filter(
        date=today,
        check_out__isnull=True
    ).count()

    absent = total_students - present_today

    return Response({
        "total_students": total_students,
        "present_today": present_today,
        "checked_out": checked_out,
        "in_center": in_center,
        "absent": absent
    })


@api_view(["GET"])
def student_stats_api(request):

    total_days = Attendance.objects.values(
        "date"
    ).distinct().count()

    data = []

    for student in Student.objects.all():

        attendance_count = Attendance.objects.filter(
            student=student
        ).count()

        attendance_percent = 0

        if total_days > 0:
            attendance_percent = round(
                (attendance_count / total_days) * 100,
                1
            )

        data.append({
            "student_id": student.student_id,
            "name": f"{student.first_name} {student.last_name}",
            "attendance_percent": attendance_percent
        })
        
        data.sort(
            key=lambda x: x["attendance_percent"],
            reverse=True
        )

    return Response(data)



@api_view(["GET"])
def attendance_trend_api(request):

    trend = (

        Attendance.objects

        .values("date")

        .annotate(
            count=Count("id")
        )

        .order_by("date")

    )

    return Response(trend)


@api_view(["GET"])
def export_attendance_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Attendance Report"

    headers = [
        "Student ID",
        "Student Name",
        "Date",
        "Check In",
        "Check Out"
    ]

    for col_num, header in enumerate(headers, 1):

        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    attendances = Attendance.objects.select_related(
        "student"
    ).order_by("-date")

    row_num = 2

    for attendance in attendances:

        sheet.cell(
            row=row_num,
            column=1
        ).value = attendance.student.student_id

        sheet.cell(
            row=row_num,
            column=2
        ).value = (
            f"{attendance.student.first_name} "
            f"{attendance.student.last_name}"
        )

        sheet.cell(
            row=row_num,
            column=3
        ).value = attendance.date

        sheet.cell(
            row=row_num,
            column=4
        ).value = (
            attendance.check_in.strftime("%I:%M %p")
            if attendance.check_in
            else ""
        )

        sheet.cell(
            row=row_num,
            column=5
        ).value = (
            attendance.check_out.strftime("%I:%M %p")
            if attendance.check_out
            else ""
        )

        row_num += 1

    response = HttpResponse(

        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response["Content-Disposition"] = (
        'attachment; filename="attendance_report.xlsx"'
    )

    workbook.save(response)

    return response

@api_view(["GET"])
def export_students_excel(request):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Students"

    headers = [

        "Student ID",

        "First Name",

        "Last Name",

        "Class",

        "Parent",

        "Phone",

        "Email"

    ]

    for col_num, header in enumerate(headers, 1):

        sheet.cell(
            row=1,
            column=col_num
        ).value = header

    students = Student.objects.all().order_by(
        "student_id"
    )

    row_num = 2

    for student in students:

        link = StudentParent.objects.filter(
            student=student
        ).select_related(
            "parent"
        ).first()

        parent_name = ""
        parent_phone = ""
        parent_email = ""

        if link:

            parent_name = (
                link.parent.full_name
            )

            parent_phone = (
                link.parent.phone_number
                if hasattr(
                    link.parent,
                    "phone_number"
                )
                else ""
            )

            parent_email = (
                link.parent.email
            )

        sheet.cell(
            row=row_num,
            column=1
        ).value = student.student_id

        sheet.cell(
            row=row_num,
            column=2
        ).value = student.first_name

        sheet.cell(
            row=row_num,
            column=3
        ).value = student.last_name

        sheet.cell(
            row=row_num,
            column=4
        ).value = (
            student.class_name
            if hasattr(
                student,
                "class_name"
            )
            else ""
        )

        sheet.cell(
            row=row_num,
            column=5
        ).value = parent_name

        sheet.cell(
            row=row_num,
            column=6
        ).value = parent_phone

        sheet.cell(
            row=row_num,
            column=7
        ).value = parent_email

        row_num += 1

    response = HttpResponse(

        content_type=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="students_report.xlsx"'
    )

    workbook.save(response)

    return response

@api_view(["GET"])
def export_attendance_csv(request):

    response = HttpResponse(
        content_type="text/csv"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; filename="attendance_report.csv"'
    )

    writer = csv.writer(response)

    writer.writerow([

        "Student ID",

        "Student Name",

        "Date",

        "Check In",

        "Check Out"

    ])

    attendances = Attendance.objects.select_related(
        "student"
    ).order_by("-date")

    for attendance in attendances:

        writer.writerow([

            attendance.student.student_id,

            f"{attendance.student.first_name} "
            f"{attendance.student.last_name}",

            attendance.date,

            attendance.check_in,

            attendance.check_out

        ])

    return response