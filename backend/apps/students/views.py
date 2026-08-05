import os
import tempfile
import zipfile
from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.db.models import Q
from django.http import FileResponse, Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas

from apps.attendance.models import Attendance

from .admin_auth import admin_required
from .forms import StudentForm, StudentImportForm
from .id_card_generator import (
    BULK_ID_CARD_PAGE_SIZE,
    ID_CARDS_PER_BULK_PAGE,
    ID_CARD_SIZE,
    draw_student_id_card,
    get_bulk_id_card_position,
    get_id_card_settings,
)
from .models import Student
from .qr_utils import (
    ensure_student_qr_code,
    get_existing_file_path,
    regenerate_student_qr_code,
)
from .services import RegistrationIntegrationService


@admin_required
def student_detail(request, pk):
    student = get_object_or_404(
        Student,
        pk=pk,
    )
    ensure_student_qr_code(student)

    attendance_history = Attendance.objects.filter(student=student).order_by("-date")

    context = {
        "student": student,
        "attendance_history": attendance_history,
    }

    return render(
        request,
        "attendance/detail.html",
        context,
    )


@admin_required
def student_qr_code(request, pk):
    student = get_object_or_404(
        Student,
        pk=pk,
    )
    ensure_student_qr_code(student)

    try:
        response = FileResponse(
            student.qr_code.open("rb"),
            content_type="image/png",
        )
    except (FileNotFoundError, OSError, ValueError):
        regenerate_student_qr_code(student)

        try:
            response = FileResponse(
                student.qr_code.open("rb"),
                content_type="image/png",
            )
        except (FileNotFoundError, OSError, ValueError) as exc:
            raise Http404("QR code is not available.") from exc

    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"

    return response


@admin_required
def student_id_card(request, pk):
    student = get_object_or_404(
        Student,
        pk=pk,
    )

    buffer = BytesIO()
    pdf = canvas.Canvas(
        buffer,
        pagesize=ID_CARD_SIZE,
    )
    draw_student_id_card(
        pdf,
        student,
        get_id_card_settings(),
    )

    pdf.save()
    buffer.seek(0)

    return FileResponse(
        buffer,
        as_attachment=True,
        filename=f"{student.student_id}.pdf",
    )


def parent_portal(request, token):
    student = get_object_or_404(
        Student,
        portal_token=token,
    )

    today = timezone.localdate()

    attendance = Attendance.objects.filter(
        student=student,
        date=today,
    ).first()

    history = Attendance.objects.filter(student=student).order_by("-date")[:30]

    context = {
        "student": student,
        "attendance": attendance,
        "history": history,
    }

    return render(
        request,
        "students/parent_portal.html",
        context,
    )


@admin_required
def import_students(request):
    form = StudentImportForm()

    if request.method == "POST":
        form = StudentImportForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            workbook = load_workbook(excel_file)
            sheet = workbook.active

            success_count = 0
            failed_count = 0
            errors = []

            for row_number, row in enumerate(
                sheet.iter_rows(
                    min_row=2,
                    values_only=True,
                ),
                start=2,
            ):
                first_name = row[0]
                last_name = row[1]
                date_of_birth = row[2]
                gender = row[3]
                class_name = row[4]
                parent_title = row[5] if len(row) > 5 else ""
                parent_name = row[6] if len(row) > 6 else ""
                parent_email = row[7] if len(row) > 7 else ""
                parent_phone = row[8] if len(row) > 8 else ""
                parent_whatsapp = row[9] if len(row) > 9 else ""
                relationship = row[10] if len(row) > 10 and row[10] else "Guardian"

                if not first_name:
                    continue

                payload = {
                    "first_name": first_name,
                    "last_name": last_name or "",
                    "date_of_birth": date_of_birth,
                    "gender": gender or "",
                    "class_name": class_name or "",
                    "parent_title": parent_title or "",
                    "parent_name": parent_name or "",
                    "parent_email": parent_email or "",
                    "parent_phone": parent_phone or "",
                    "parent_whatsapp": parent_whatsapp or "",
                    "relationship": relationship,
                    "photo": None,
                    "_base_url": request.build_absolute_uri("/"),
                }

                try:
                    RegistrationIntegrationService.register(payload)
                    success_count += 1
                except Exception as exc:
                    failed_count += 1
                    errors.append(f"Row {row_number}: {exc}")

            if success_count:
                messages.success(
                    request,
                    f"{success_count} student(s) imported successfully.",
                )

            if failed_count:
                messages.warning(
                    request,
                    f"{failed_count} row(s) failed during import.",
                )

                for error in errors[:20]:
                    messages.error(
                        request,
                        error,
                    )

            return redirect("students:student_list")

    context = {
        "form": form,
    }

    return render(
        request,
        "students/import.html",
        context,
    )


@admin_required
def create_student(request):
    form = StudentForm()

    if request.method == "POST":
        form = StudentForm(
            request.POST,
            request.FILES,
        )

        if form.is_valid():
            teaching_class = form.cleaned_data.get("teaching_class")

            payload = {
                "first_name": form.cleaned_data["first_name"],
                "last_name": form.cleaned_data.get("last_name", ""),
                "date_of_birth": form.cleaned_data.get("date_of_birth"),
                "gender": form.cleaned_data.get("gender", ""),
                "class_name": (
                    teaching_class.name
                    if teaching_class
                    else form.cleaned_data["class_name"]
                ),
                "teaching_class": teaching_class,
                "parent_title": form.cleaned_data.get("parent_title", ""),
                "parent_name": form.cleaned_data.get("parent_name", ""),
                "parent_email": form.cleaned_data.get("parent_email", ""),
                "parent_phone": form.cleaned_data.get("parent_phone", ""),
                "parent_whatsapp": form.cleaned_data.get("parent_whatsapp", ""),
                "relationship": form.cleaned_data.get("relationship", "Guardian"),
                "photo": request.FILES.get("photo") or form.cleaned_data.get("photo"),
                "_base_url": request.build_absolute_uri("/"),
            }

            try:
                result = RegistrationIntegrationService.register(payload)
                student = result["student"]

                if result["parent_created"]:
                    messages.success(
                        request,
                        (
                            "Student registered successfully. "
                            f"Temporary parent password: {result['temporary_password']}"
                        ),
                    )
                else:
                    messages.success(
                        request,
                        f"{student.first_name} registered successfully.",
                    )

                return redirect(
                    "students:student_detail",
                    pk=student.pk,
                )

            except Exception as exc:
                messages.error(
                    request,
                    str(exc),
                )

    context = {
        "form": form,
    }

    return render(
        request,
        "students/create.html",
        context,
    )


@admin_required
def edit_student(request, pk):
    student = get_object_or_404(
        Student,
        pk=pk,
    )

    form = StudentForm(
        instance=student,
    )

    if request.method == "POST":
        form = StudentForm(
            request.POST,
            request.FILES,
            instance=student,
        )

        if form.is_valid():
            student = form.save()

            RegistrationIntegrationService.sync_parent_for_student(
                student=student,
                data={
                    "parent_title": form.cleaned_data.get("parent_title", ""),
                    "parent_name": form.cleaned_data.get("parent_name", ""),
                    "parent_email": form.cleaned_data.get("parent_email", ""),
                    "parent_phone": form.cleaned_data.get("parent_phone", ""),
                    "parent_whatsapp": form.cleaned_data.get("parent_whatsapp", ""),
                    "relationship": form.cleaned_data.get("relationship", ""),
                },
            )

            messages.success(
                request,
                "Student updated successfully.",
            )

            return redirect(
                "students:student_detail",
                pk=student.pk,
            )

    context = {
        "form": form,
        "student": student,
    }

    return render(
        request,
        "students/edit.html",
        context,
    )


@admin_required
def delete_student(request, pk):
    student = get_object_or_404(
        Student,
        pk=pk,
    )

    if request.method == "POST":
        student_name = f"{student.first_name} {student.last_name}"

        student.delete()

        messages.success(
            request,
            f"{student_name} deleted successfully.",
        )

        return redirect("students:student_list")

    context = {
        "student": student,
    }

    return render(
        request,
        "students/delete.html",
        context,
    )


@admin_required
@require_POST
def regenerate_student_qr(request, pk):
    student = get_object_or_404(
        Student,
        pk=pk,
    )
    regenerate_student_qr_code(student)
    messages.success(
        request,
        f"QR code regenerated for {student.full_name}.",
    )

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")

    if next_url:
        return redirect(next_url)

    return redirect(
        "students:student_detail",
        pk=student.pk,
    )


@admin_required
@require_POST
def regenerate_missing_qr_codes(request):
    regenerated_count = 0

    for student in Student.objects.all().order_by("id"):
        if ensure_student_qr_code(student):
            regenerated_count += 1

    if regenerated_count:
        messages.success(
            request,
            f"{regenerated_count} missing QR code(s) regenerated.",
        )
    else:
        messages.success(
            request,
            "All students already have QR codes.",
        )

    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")

    if next_url:
        return redirect(next_url)

    return redirect("school_admin:student_list")


@admin_required
def bulk_qr_download(request):
    if request.method != "POST":
        return redirect("students:student_list")

    student_ids = request.POST.getlist("students")

    if not student_ids:
        messages.warning(
            request,
            "No students were selected.",
        )

        return redirect("students:student_list")

    students = Student.objects.filter(id__in=student_ids).order_by("student_id")

    buffer = BytesIO()

    with zipfile.ZipFile(
        buffer,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
    ) as archive:
        for student in students:
            ensure_student_qr_code(student)
            qr_code_path = get_existing_file_path(student.qr_code)

            if qr_code_path and os.path.exists(qr_code_path):
                archive.write(
                    qr_code_path,
                    arcname=f"{student.student_id}.png",
                )

    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/zip",
    )
    response["Content-Disposition"] = (
        'attachment; filename="student_qr_codes.zip"'
    )

    return response


@admin_required
def bulk_id_cards(request):
    if request.method != "POST":
        return redirect("students:student_list")

    student_ids = request.POST.getlist("students")
    download_all = request.POST.get("all_students") == "1"

    if not student_ids and not download_all:
        messages.warning(
            request,
            "No students were selected.",
        )

        next_url = request.POST.get("next") or request.META.get("HTTP_REFERER")

        if next_url:
            return redirect(next_url)

        return redirect("students:student_list")

    students = Student.objects.all()

    if not download_all:
        students = students.filter(id__in=student_ids)

    students = students.order_by("student_id")

    output = tempfile.TemporaryFile()
    pdf = canvas.Canvas(
        output,
        pagesize=BULK_ID_CARD_PAGE_SIZE,
    )
    id_card_settings = get_id_card_settings()

    for index, student in enumerate(students.iterator(chunk_size=100)):
        if index and index % ID_CARDS_PER_BULK_PAGE == 0:
            pdf.showPage()

        pdf.setPageSize(BULK_ID_CARD_PAGE_SIZE)
        x, y = get_bulk_id_card_position(index)
        draw_student_id_card(
            pdf,
            student,
            id_card_settings,
            x=x,
            y=y,
        )

    pdf.save()
    output.seek(0)

    return FileResponse(
        output,
        as_attachment=True,
        filename="student_id_cards_a4.pdf",
    )


@admin_required
def download_student_template(request):
    file_path = os.path.join(
        settings.BASE_DIR,
        "static",
        "templates",
        "student_import_template.xlsx",
    )

    return FileResponse(
        open(file_path, "rb"),
        as_attachment=True,
        filename="student_import_template.xlsx",
    )


@admin_required
def download_student_template(request):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"

    headers = [
        "first_name",
        "last_name",
        "date_of_birth",
        "gender",
        "class_name",
        "parent_title",
        "parent_name",
        "parent_email",
        "parent_phone",
        "parent_whatsapp",
        "relationship",
    ]

    header_fill = PatternFill(
        start_color="1E40AF",
        end_color="1E40AF",
        fill_type="solid",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(
            row=1,
            column=col_num,
        )
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font

    sample_row = [
        "John",
        "Doe",
        "2015-04-20",
        "Male",
        "JSS1",
        "Mrs",
        "Jane Doe",
        "jane@example.com",
        "08012345678",
        "08012345678",
        "Mother",
    ]

    for col_num, value in enumerate(sample_row, start=1):
        sheet.cell(
            row=2,
            column=col_num,
        ).value = value

    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            value = str(cell.value or "")
            if len(value) > max_length:
                max_length = len(value)

        sheet.column_dimensions[column_letter].width = max_length + 5

    sheet.freeze_panes = "A2"

    instructions = workbook.create_sheet(title="Instructions")

    instructions["A1"] = (
        "CODECAMP ATTENDANCE SYSTEM STUDENT IMPORT TEMPLATE"
    )
    instructions["A1"].font = Font(
        bold=True,
        size=14,
    )

    instructions["A3"] = "Fill one student per row."
    instructions["A4"] = "Do not change the header names."
    instructions["A5"] = "Date format must be YYYY-MM-DD."
    instructions["A6"] = "Gender examples: Male, Female."
    instructions["A7"] = "Relationship examples: Father, Mother, Guardian."
    instructions["A8"] = "Parent email is optional."
    instructions["A9"] = "Parent phone numbers should include leading zero."

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = (
        'attachment; filename="student_import_template.xlsx"'
    )

    workbook.save(response)

    return response


@admin_required
def student_list(request):
    search = request.GET.get("search", "").strip()

    students = Student.objects.all().order_by("-id")

    if search:
        students = students.filter(
            Q(student_id__icontains=search)
            | Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(parent_name__icontains=search)
        )

    context = {
        "students": students,
        "search": search,
        "total_students": students.count(),
    }

    return render(
        request,
        "students/list.html",
        context,
    )
